from openpyxl.styles import Font

from CbsObjects.CbsLink import CbsLink
from CbsObjects.Pages.SubjectPage import SubjectPage
from enum import Enum

import sys
import json
import sqlite3

####,encoding="utf-8"
from CbsObjects.TestDetails import TestDetails

ROOT_PATH = sys.path[1]

ERROR_TYPE_DIC = None
WEP_PART_TYPE_DIC = None
HE_SUBJECT_PAGES_DIC = None
TESTS_DETAILS_DIC = None


class DB:
    def __init__(self):
        self.path = ROOT_PATH + "\\DataBase\\MainDB"
        self.__db = sqlite3.connect(self.path, check_same_thread=False)
        self.__cursor = self.__db.cursor()

    def load_en_subject_pages_dic(self):
        self.__cursor.execute("SELECT * FROM PAGES_DIC WHERE lang='EN' ")
        data = self.__cursor.fetchall()
        return data

    def save_test_results(self, errors):
        insert = "INSERT INTO TEST_RESULTS VALUES (?,?,?,?,?);"
        for i in errors:
            self.__cursor.execute(insert, i.to_list())
        self.__db.commit()
        print('--------------------------------------------------------------')
        print('new {} errors was added successfully to db'.format(len(errors)))
        print('--------------------------------------------------------------')

    def load_test_data(self, test_key):
        select = "SELECT * FROM TEST_RESULTS WHERE test_id=? "
        self.__cursor.execute(select, (test_key,))
        data = self.__cursor.fetchall()
        return data

    def add_new_test(self, details: TestDetails):
        insert = "INSERT INTO TEST_DETAILS VALUES (?,?,?,?,?,?,?);"
        self.__cursor.execute(insert, details.to_list())
        self.__db.commit()

    def __update_new_pages(self, pages):

        index = 400
        insert = "INSERT INTO PAGES_DIC VALUES (?,?,?,?);"

        for page in pages:
            p = [page.name, page.url, index, 'EN']
            if not p[0]:
                p[0] = list(p[1].split('/'))[-1].split('.')[0]
            # en_pages.append(p)
            index += 1
            try:
                self.__cursor.execute(insert, p)
                self.__db.commit()
            except Exception as e:
                print(e)
                print('not inserted: ', p[0], p[1])

    def load_tests_details_dic(self):
        self.__cursor.execute("SELECT * FROM TEST_DETAILS")
        data = self.__cursor.fetchall()

        return data

    def load_errors_dic(self):
        self.__cursor.execute("SELECT * FROM ERRORS_DIC")
        data = self.__cursor.fetchall()
        return data

    def load_web_parts_dic(self):
        self.__cursor.execute("SELECT * FROM WEB_PARTS_DIC")
        data = self.__cursor.fetchall()
        return data

    def load_he_subject_pages_dic(self):
        self.__cursor.execute("SELECT * FROM PAGES_DIC WHERE lang='HE' ")
        data = self.__cursor.fetchall()
        # return data[0:6]
        return data

    def load_he_subject_pages_dic_temp(self):
        self.__cursor.execute("SELECT * FROM PAGES_DIC WHERE lang='HE' ")
        data = self.__cursor.fetchall()
        return data

    def __del__(self):
        self.__cursor.close()
        self.__db.close()

    def update_test_details(self, details: TestDetails):
        select = "SELECT * FROM TEST_DETAILS WHERE test_id=?"
        self.__cursor.execute(select,(details.key(),))
        test_row = self.__cursor.fetchall()
        if len(test_row):  # if test details is in database
            update = '''UPDATE TEST_DETAILS 
            SET
            end_date=?,
            end_time=?,
            scanned=?
            WHERE
            test_id = ?;'''
            self.__cursor.execute(update, (str(details.end_date()), str(details.end_time()), str(details.scanned()),str(details.key())))
            self.__db.commit()
        else:
            raise Exception('There is no test-id {} in database, can not update test details'.format(details.key()))


    def exist(self):
        select = "SELECT * FROM TEST_DETAILS WHERE test_id='00f80ee7-1ff3-45f4-80e0-94fd70404a27'"
        self.__cursor.execute(select)
        details = self.__cursor.fetchall()
        if len(details):  # update candidates
            scanned: str = details[0][6]
            str_ids = scanned.replace('}', '').replace('{', '').split(',')
            s = map(lambda id:int(id),str_ids)
            r = set(s)
            print(r)


db = DB()


class DataBase:
    @classmethod
    def get_CBS_he_pages(cls):
        pages = []
        try:
            raw_pages = db.load_he_subject_pages_dic()
            for page in raw_pages:
                link = CbsLink(page_name=page[0], url=page[1])
                pages.append(SubjectPage(link, page[0], id=page[2], lang=page[3]))
        except Exception as e:
            print('database exception, can not load pages dict', e)
            raise e
        pages = list(set(pages))
        excluded = ('/search/', '/Surveys/', '/Documents/', '/publications/')
        pages = list(filter(lambda x: all(s not in x.link.url for s in excluded), pages))
        pages.sort(key=lambda x: x.name)
        return pages

    @classmethod
    def get_CBS_en_pages(cls):
        pages = []
        try:
            raw_pages = db.load_en_subject_pages_dic()
            for page in raw_pages:
                link = CbsLink(page_name=page[0], url=page[1])
                pages.append(SubjectPage(link, page[0], id=page[2], lang=page[3]))
        except Exception as e:
            print('database exception, can not load pages dict', e)
            raise e
        pages = list(set(pages))
        excluded = ('/search/', '/Surveys/', '/Documents/', '/publications/')
        pages = list(filter(lambda x: all(s not in x.link.url for s in excluded), pages))
        pages.sort(key=lambda x: x.name)
        return pages

    @classmethod
    def save_test_results(cls, errors):
        try:
            db.save_test_results(errors)
        except Exception as e:
            print('exception in db new db insertion')
            raise e

    @classmethod
    def get_test_results(cls, test_key):
        try:
            return db.load_test_data(test_key)
        except Exception as e:
            print("exception in db, couldn't load test results")
            raise e

    @classmethod
    def load_xpath(cls, key):
        with open(ROOT_PATH + '\\Configuration\\xpath.json', 'rb') as f:
            data = json.load(f)
            f.close()
        return data['XPath'][0][key]

    @classmethod
    def get_webdriver_path(cls):
        with open(ROOT_PATH + '\\Configuration\\webdriver_path.json', 'rb') as file:
            data = json.load(file)
            file.close()
        return data['driver_path']

    @classmethod
    def get_pdf_test_results(cls, test_ID):
        data = cls.get_test_results(test_ID)
        return Converter.to_pdf(data)

    @classmethod
    def get_html_test_results(cls, test_ID):
        data = cls.get_test_results(test_ID)
        return Converter.to_html(data)

    @classmethod
    def init_new_test(cls, test_details):
        db.add_new_test(test_details)

    @classmethod
    def get_excel_test_results(cls, test_ID):
        data = cls.get_test_results(test_ID)
        return Converter.to_excel(data)

    @classmethod
    def load_wpart_dic(cls):
        try:
            w_parts = db.load_web_parts_dic()
            return dict([(wp_id, (name, real_name)) for wp_id, name, real_name in w_parts])
        except Exception as e:
            print("exception in db, couldn't load web parts dictionary")
            raise e

    @classmethod
    def load_he_pages_dic(cls):
        try:
            pages = db.load_he_subject_pages_dic()
            return dict([(p_id, (name, url, lang)) for name, url, p_id, lang in pages])
        except Exception as e:
            print("exception in db, couldn't load hebrew subject pages dictionary")
            raise e

    @classmethod
    def load_tests_details_dic(cls):
        try:
            t_details = db.load_tests_details_dic()
            return dict([(test_key, (s_date, s_time, e_date, e_time, candidates, scanned)) for
                         test_key, s_date, s_time, e_date, e_time, candidates, scanned in t_details])
        except Exception as e:
            print("exception in db, couldn't load tests details dictionary")
            raise e

    @classmethod
    def load_error_dic(cls):
        try:
            errors = db.load_errors_dic()
            return dict([(e_id, real_name) for e_id, name, real_name in errors])
        except Exception as e:
            print("exception in db, couldn't load errors details dictionary")
            raise e

    @classmethod
    def add_test_details(cls, test_details):
        db.update_test_details(test_details)




class DicData(Enum):
    ERROR_TYPE_DIC = DataBase.load_error_dic()
    WEP_PART_TYPE_DIC = DataBase.load_wpart_dic()
    HE_SUBJECT_PAGES_DIC = DataBase.load_he_pages_dic()
    TESTS_DETAILS_DIC = DataBase.load_tests_details_dic()


class Links(Enum):
    CBS_HOME_PAGE_HE = 'https://www.cbs.gov.il/he/Pages/default.aspx'
    CBS_MAP_SITE_HE = 'https://www.cbs.gov.il/he/pages/sitemap.aspx'
    CBS_MAP_SITE_EN = 'https://www.cbs.gov.il/en/Pages/sitemap.aspx'
    ROOT_XPATH = DataBase.load_xpath('ROOT_XPATH')
    ROOT_DIR = sys.path[1]
    CHROME_DRIVER = ROOT_DIR + "/" + DataBase.get_webdriver_path()
    HEBREW_STATS_XPATH = DataBase.load_xpath('HEBREW_STATS_XPATH')
    RIGHT_EXTRA_PARTS_XPATH = DataBase.load_xpath('RIGHT_EXTRA_PARTS_XPATH')
    LEFT_EXTRA_PARTS_XPATH = DataBase.load_xpath('LEFT_EXTRA_PARTS_XPATH')
    TOOLS_AND_DB_XPATH = DataBase.load_xpath('TOOLS_AND_DB_XPATH')
    SUMMARY_XPATH = DataBase.load_xpath('SUMMARY_XPATH')
    TOP_BOX_XPATH = DataBase.load_xpath('TOP_BOX_XPATH')
    SUB_SUBJECTS_XPATH = DataBase.load_xpath('SUB_SUBJECTS_XPATH')
    PRESS_RELEASES_XPATH = DataBase.load_xpath('PRESS_RELEASES_XPATH')
    TABLES_AND_CHARTS_XPATH = DataBase.load_xpath('TABLES_AND_CHARTS_XPATH')
    PUBLICATIONS_XPATH = DataBase.load_xpath('PUBLICATIONS_XPATH')
    GEOGRAPHIC_ZONE_XPATH = DataBase.load_xpath('GEOGRAPHIC_ZONE_XPATH')
    INTERNATIONAL_COMPARISONS_XPATH = DataBase.load_xpath('INTERNATIONAL_COMPARISONS_XPATH')  # new
    MORE_LINKS_XPATH = DataBase.load_xpath('MORE_LINKS_XPATH')  # new
    CONFERENCES_AND_SEMINARS_XPATH = DataBase.load_xpath('CONFERENCES_AND_SEMINARS_XPATH')  # new
    VIDEOS_LINKS_XPATH = DataBase.load_xpath('VIDEOS_LINKS_XPATH')  # new
    PICTURES_LINKS_XPATH = DataBase.load_xpath('PICTURES_LINKS_XPATH')  # new
    PRESENTATIONS_XPATH = DataBase.load_xpath('PRESENTATIONS_XPATH')  # new


import itertools
import operator
from openpyxl import Workbook
import pdfkit
from CbsObjects.Error import Error

PDF_PATH = ROOT_PATH + '\\Resources\\wkhtmltopdf\\bin\\wkhtmltopdf.exe'
PDF_CONFIG = pdfkit.configuration(wkhtmltopdf=PDF_PATH)
TEMPLATE = '''<!DOCTYPE html>
<html lang="en">
<head>
    <meta charset="UTF-8">
    <title>Title</title>
</head>
<body>
'''


class Converter:
    @classmethod
    def error_to_short_str(cls, error: Error):
        details = cls.error_to_details(error)
        result = str(details[2]) + ': ' + str(details[3])
        if details[4]:
            result = result + ' at object {}'.format(details[4])
        return result

    @classmethod
    def error_to_details(cls, error: Error):
        page_name = DicData.HE_SUBJECT_PAGES_DIC.value.get(error.page_id)[0]
        web_part_description = DicData.WEP_PART_TYPE_DIC.value.get(error.wp_type)[1]
        error_description = DicData.ERROR_TYPE_DIC.value.get(error.type)
        return [error.test_id, page_name, web_part_description, error_description, error.index]

    @classmethod
    def tuple_error_to_details(cls, err):
        page_name = DicData.HE_SUBJECT_PAGES_DIC.value.get(err[1])[0]
        web_part_description = DicData.WEP_PART_TYPE_DIC.value.get(err[2])[1]
        error_description = DicData.ERROR_TYPE_DIC.value.get(err[3])
        return [err[0], page_name, web_part_description, error_description, err[4]]

    @classmethod
    def to_html(cls, errors: []):
        if not errors:
            return None
        data = TEMPLATE + cls.__html_report_details()

        it_by_test_id = itertools.groupby(errors, operator.itemgetter(0))
        for test_id, errors_by_test_id in it_by_test_id:
            data = data + cls.__html_test_details(test_id)
            it_by_page_id = itertools.groupby(errors_by_test_id, operator.itemgetter(1))
            for page_id, errors_ in it_by_page_id:
                data = data + cls.__html_page_details(page_id)
                # errors details loop
                for err in errors_:
                    data = data + cls.__html_error_details(err)

        data = data + "</body></html>"
        return data

    @classmethod
    def to_pdf(cls, errors: []):
        if not errors:
            return None
        data = TEMPLATE + cls.__html_report_details()

        it_by_test_id = itertools.groupby(errors, operator.itemgetter(0))
        for test_id, errors_by_test_id in it_by_test_id:
            data = data + cls.__html_test_details(test_id)
            it_by_page_id = itertools.groupby(errors_by_test_id, operator.itemgetter(1))
            for page_id, errors_ in it_by_page_id:
                data = data + cls.__html_page_details(page_id)
                # errors details loop
                for err in errors_:
                    data = data + cls.__html_error_details(err)
        # data = TEMPLATE + '''<p class="Title">
        #                         Test started on: {}<br>
        #                         Test ended on: {}<br>
        #                         Candidates: {}<br>
        #                         Scanned pages: {}<br>
        #                         Test ID: {}<br></p>'''.format(t.start_date_str(), t.end_date_str(), len(t.candidates()), len(t.scanned()), t.ID())
        data = data + "</body></html>"
        data = pdfkit.from_string(data, configuration=PDF_CONFIG)
        return data

    @classmethod
    def to_excel(cls, errors: []):
        if not errors:
            return None

        workbook = Workbook()

        sheet = workbook.active
        sheet.cell(row=1,column=1).value = 'PAGE'
        sheet.cell(row=1,column=1).font = Font(size=18,bold=True)
        sheet.cell(row=1,column=2).value = 'WEB PART'
        sheet.cell(row=1,column=2).font = Font(size=18,bold=True)
        sheet.cell(row=1,column=3).value = 'DESCRIPTION'
        sheet.cell(row=1,column=3).font = Font(size=18,bold=True)
        sheet.cell(row=1,column=4).value = 'LINE / LOCATION'
        sheet.cell(row=1,column=4).font = Font(size=18,bold=True)

        for e in errors:
            err = cls.error_tuple_to_details(e)
            sheet.append((err[1],err[2],err[3],err[4]))
        # filename=ROOT_PATH + '\\DataBase\\{}'.format("sample_test.xlsx")
        return workbook.save()
        # return workbook

    @classmethod
    def to_excel_temp(cls, errors: []):
        if not errors:
            return None

        workbook = Workbook()

        sheet = workbook.active
        sheet.merge_cells('A1:H4')
        sheet.cell(row=1, column=1).value = cls.__report_details()
        current_row = 5
        it_by_test_id = itertools.groupby(errors, operator.itemgetter(0))
        for test_id, errors_by_test_id in it_by_test_id:
            # test id value
            sheet.merge_cells('A{}:F{}'.format(current_row, current_row + 2))
            top_left_cell = sheet['A{}'.format(current_row)]
            top_left_cell.value = cls.__test_details(test_id)
            # sheet.cell(current_row, column=1).value = cls.__test_details(test_id)
            current_row += 3

            it_by_page_id = itertools.groupby(errors_by_test_id, operator.itemgetter(1))
            for page_id, errors_ in it_by_page_id:
                # page id value
                sheet.merge_cells('A{}:F{}'.format(current_row, current_row + 1))
                top_left_cell = sheet['A{}'.format(current_row)]
                top_left_cell.value = cls.__page_details(page_id)
                # sheet.cell(current_row, column=1).value = cls.__page_details(page_id)
                current_row += 2

                # errors details loop
                for err in errors_:
                    print(err)
                    # sheet.append(err)
                    top_left_cell = sheet['A{}'.format(current_row)]
                    top_left_cell.value = str(err)
                    current_row += 1
        workbook.save(filename=ROOT_PATH + '\\DataBase\\{}'.format("sample.xlsx"))

    @classmethod
    def __html_test_details(cls, test_id):
        return '<h2>{}</h2>\n'.format(test_id)

    @classmethod
    def __html_page_details(cls, page_id):
        page_details = Converter.page_details(page_id)
        return '<h3><a href={}>{}</a></h3>\n'.format(page_details[1],page_details[0])

    @classmethod
    def __html_error_details(cls, err):
        e = Converter.error_tuple_to_short_str(err)
        return '<h4>{}</h4>\n'.format(e)

    @classmethod
    def __html_report_details(cls):
        return '<h1>Report details </h1>\n'

    @classmethod
    def __report_details(cls):
        return 'Report details'

    @classmethod
    def __page_details(cls, page_id):
        return str(page_id)

    @classmethod
    def __test_details(cls, test_id):
        return str(test_id)

    @classmethod
    def page_details(cls, page_id):
        return DicData.HE_SUBJECT_PAGES_DIC.value.get(page_id)

    @classmethod
    def error_tuple_to_short_str(cls, err):
        details = cls.tuple_error_to_details(err)
        result = str(details[2]) + ': ' + str(details[3])
        if details[4]:
            result = result + ' at object {}'.format(details[4])
        return result

    @classmethod
    def error_tuple_to_details(cls, e):
        if not DicData.HE_SUBJECT_PAGES_DIC.value.get(e[1]):
            print('none type in: ', e[1])
            print(DicData.HE_SUBJECT_PAGES_DIC.value)
            return
        print(type(DicData.HE_SUBJECT_PAGES_DIC.value.get(e[1])))
        page_name = DicData.HE_SUBJECT_PAGES_DIC.value.get(e[1])[0]
        web_part_description = DicData.WEP_PART_TYPE_DIC.value.get(e[2])[1]
        error_description = DicData.ERROR_TYPE_DIC.value.get(e[3])
        return [e[0], page_name, web_part_description, error_description, e[4]]


db.exist()

DataBase.get_excel_test_results('''0b0920b3-088a-43ab-9a6b-cbdf8bf2e293''')
# print(DataBase.load_wpart_dic())
# #######
# links = list(set(DataBase.get_CBS_en_links()))
#
# print(links)
# db.update_new_pages(links)
###########

# st =  'https://www.cbs.gov.il/en/subjects/Pages/Index-of-Compactness-of-Municipalities-and-Local-Councils.aspx'
# print(st.split('/')[-1].split('.')[0])
# x = DB()
# # print(x.get_he_subject_pages())
# x.add_new_test([0, 1, 2, 3])
# print(sys.path[1]+"\\DL\\DB\\")
# driver = TestUtility.get_sessions()[0]
# driver.get('https://getsharex.com/')
#
# print(Links.ROOT_XPATH.value)
# summ = [1,2,3,4,5]
# path = '02_Jun_2021_10.48.50'
# DataBase.save_summary_result(file_key=path,summery=summ)
# print(DataBase.get_CBS_he_pages()[30])


# @classmethod
# def save_test_result(cls, test_key, page: SubjectPage):
#     try:
#         path = ROOT_PATH + '\\TestData\\logs'
#         file = path + '\\' + test_key + '.html'
#         with open(file, 'a', encoding='utf-8') as f:
#             style = 'style={color:red; font-size: large; }'
#             page_link = '<h1 {}><a style="color:red" href="{}" target="_blank" >{}</a></h1><br>'.format(style,
#                                                                                                         page.link.url,
#                                                                                                         page.name)
#             errors = ('<h1 {}>' + str(page.error_to_str()) + '</h1><br>').format(style)
#             f.write(page_link + errors)
#         f.close()
#     except Exception as e:
#         print('exception in db')
#         raise e

# @classmethod
#     def save_summary_result(cls, file_key, summary):
#         sum = '<h1 style="color:black" style={color:red; font-size: large; }>Test started on: ' + str(
#             summary[0]) + ' ' + str(summary[1]) + '<br>'
#         sum += 'Total pages: ' + str(summary[2]) + '<br>'
#         sum += 'Tested: ' + str(summary[3]) + '<br>'
#         sum += 'Total error pages: ' + str(summary[4]) + '</h1>'
#         file_name = file_key
#         head = '''<!DOCTYPE html>
#         <html lang="en">
#         <head>
#             <meta charset="UTF-8">
#             <title>Title</title>
#         </head>
#         <body>'''
#         tail = '</body></html>'
#
#         try:
#             path = ROOT_PATH + '\\TestData\\logs'
#             file = path + '\\' + file_name + '.html'
#             with open(file, 'r', encoding='utf-8') as f:
#                 content = f.read()
#                 f.close()
#             with open(file, 'w', encoding='utf-8') as f:
#                 f.write(head + sum + '<br>' + content + tail)
#             f.close()
#         except Exception as e:
#             print('exception in db writing summery')
#             raise e

#     scanned_as_str = details.to_list()[]
        #     scanned_list = scanned_as_str.replace('}', '').replace('{', '').split(',')
        #     scanned_as_int = map(lambda id: int(id), scanned_list)
        #     scanned_set = set(scanned_as_int)
        #     scanned_set.
        #
        # select = "SELECT * FROM TEST_DETAILS WHERE test_id='?'"
        # self.__cursor.execute(select, details.key())
        # details = self.__cursor.fetchall()
        # if len(details):  # update candidates
        #     scanned = details[0][6]
        #     scanned = set(scanned)
        #
        # exist = '''IF EXISTS ( SELECT 1 FROM TEST_DETAILS WHERE test_id = '?')
        #         BEGIN
        #             SELECT * FROM TEST_DETAILS WHERE test_id='? '
        #         END'''
        # insert = "INSERT INTO TEST_DETAILS VALUES (?,?,?,?,?,?,?);"
        # self.__cursor.execute(insert, details.to_list())
        # self.__db.commit()