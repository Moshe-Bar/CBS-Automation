import itertools
import operator
import sys

from openpyxl import Workbook

import pdfkit
from CbsObjects.Error import Error


ROOT_PATH = sys.path[1]
PDF_PATH = ROOT_PATH + '\\Resources\\wkhtmltopdf\\bin\\wkhtmltopdf.exe'
PDF_CONFIG = pdfkit.configuration(wkhtmltopdf=PDF_PATH)
TEMPLATE = '''<!DOCTYPE html>
<html lang="en">
<head>
    <meta charset="UTF-8">
    <title>Title</title>
</head>
<body>
'''


class Converter:
    @classmethod
    def error_to_details(cls, error: Error):
        code_list = error.to_list()
        # error_details = DicData.ERROR_TYPE_DIC.value

    @classmethod
    def to_html(cls, errors: []):
        pass

    @classmethod
    def to_pdf(cls, errors: []):
        if not errors:
            return None
        data = TEMPLATE + cls.__html_report_details()

        it_by_test_id = itertools.groupby(errors, operator.itemgetter(0))
        for test_id, errors_by_test_id in it_by_test_id:
            data = data + cls.__html_test_details(test_id)
            it_by_page_id = itertools.groupby(errors_by_test_id, operator.itemgetter(1))
            for page_id, errors_ in it_by_page_id:
                data = data + cls.__html_page_details(page_id)
                # errors details loop
                for err in errors_:
                    data = data + cls.__html_error_details(err)
        # data = TEMPLATE + '''<p class="Title">
        #                         Test started on: {}<br>
        #                         Test ended on: {}<br>
        #                         Candidates: {}<br>
        #                         Scanned pages: {}<br>
        #                         Test ID: {}<br></p>'''.format(t.start_date_str(), t.end_date_str(), len(t.candidates()), len(t.scanned()), t.ID())
        data = data + "</body></html>"
        data = pdfkit.from_string(data, configuration=PDF_CONFIG)
        return data

    @classmethod
    def to_excel(cls, errors: []):
        if not errors:
            return None

        workbook = Workbook()

        sheet = workbook.active
        sheet.merge_cells('A1:H4')
        sheet.cell(row=1, column=1).value = cls.__report_details()
        current_row = 5
        it_by_test_id = itertools.groupby(errors, operator.itemgetter(0))
        for test_id, errors_by_test_id in it_by_test_id:
            # test id value
            sheet.merge_cells('A{}:F{}'.format(current_row, current_row + 2))
            top_left_cell = sheet['A{}'.format(current_row)]
            top_left_cell.value = cls.__test_details(test_id)
            # sheet.cell(current_row, column=1).value = cls.__test_details(test_id)
            current_row += 3

            it_by_page_id = itertools.groupby(errors_by_test_id, operator.itemgetter(1))
            for page_id, errors_ in it_by_page_id:
                # page id value
                sheet.merge_cells('A{}:F{}'.format(current_row, current_row + 1))
                top_left_cell = sheet['A{}'.format(current_row)]
                top_left_cell.value = cls.__page_details(page_id)
                # sheet.cell(current_row, column=1).value = cls.__page_details(page_id)
                current_row += 2

                # errors details loop
                for err in errors_:
                    print(err)
                    # sheet.append(err)
                    top_left_cell = sheet['A{}'.format(current_row)]
                    top_left_cell.value = str(err)
                    current_row += 1
        workbook.save(filename=ROOT_PATH + '\\DataBase\\{}'.format("sample.xlsx"))

    @classmethod
    def __html_test_details(cls, test_id):
        return '<h2>{}</h2>\n'.format(test_id)

    @classmethod
    def __html_page_details(cls, page_id):
        return '<h3>{}</h3>\n'.format(page_id)

    @classmethod
    def __html_error_details(cls, err):
        return '<h4>{}</h4>\n'.format(err)

    @classmethod
    def __html_report_details(cls):
        return '<h1>Report details </h1>\n'

    @classmethod
    def __report_details(cls):
        return 'Report details'

    @classmethod
    def __page_details(cls, page_id):
        return str(page_id)

    @classmethod
    def __test_details(cls, test_id):
        return str(test_id)
